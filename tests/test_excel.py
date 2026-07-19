import openpyxl
from pathlib import Path

from apply_workflow.utils.excel import (
    read_applications,
    append_application,
    update_application,
    delete_application,
)


def make_wb(tmp_path: Path) -> Path:
    # Headers use different casing/spaces to exercise normalisation
    headers = [
        "Company",
        "POST",
        "Email",
        "Keywords",
        "Description",
        "HR LINkdin",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    # Valid row
    ws.append(["Acme", "Engineer", "hr@acme.com", "python, ci", "Cool role", "https://lnk" ])
    # Row missing company should be skipped by read_applications
    ws.append(["", "Designer", "hr@example.com", "figma", "UI work", ""]) 

    path = tmp_path / "apps.xlsx"
    wb.save(path)
    return path


def test_read_applications_skips_missing_company(tmp_path):
    excel_path = make_wb(tmp_path)
    apps = read_applications(str(excel_path))
    assert len(apps) == 1
    app = apps[0]
    assert app.company == "Acme"
    assert app.post == "Engineer"
    assert app.to_email == "hr@acme.com"
    assert app.keywords == "python, ci"
    assert app.description == "Cool role"


def test_append_update_delete_application(tmp_path):
    excel_path = make_wb(tmp_path)

    # Append new row (order should match headers regardless of field order provided)
    append_application(str(excel_path), {
        "company": "Globex",
        "post": "DevOps",
        "to_email": "jobs@globex.com",
        "keywords": "aws,k8s",
        "description": "SRE",
        "hr_linkedin": "https://hr.example",
    })

    # Update the just-appended row (it's at row 4: 1 header + 2 existing rows)
    update_application(str(excel_path), 4, {"post": "Senior DevOps"})

    # Verify update via openpyxl directly
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    assert ws.max_row == 4
    row4 = [cell.value for cell in ws[4]]
    # Header order is: Company, POST, Email, Keywords, Description, HR LINkdin
    assert row4[0] == "Globex"
    assert row4[1] == "Senior DevOps"
    assert row4[2] == "jobs@globex.com"

    # Delete row 2 (the original valid row)
    delete_application(str(excel_path), 2)
    assert ws.max_row == 4  # openpyxl object not auto-refreshed; reopen to verify
    wb.close()

    wb2 = openpyxl.load_workbook(excel_path)
    ws2 = wb2.active
    # Now: header + (blank-company row) + Globex
    assert ws2.max_row == 3
    row3 = [cell.value for cell in ws2[3]]
    assert row3[0] == "Globex"
    wb2.close()
