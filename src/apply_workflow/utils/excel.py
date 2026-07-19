import logging
from pathlib import Path

import openpyxl

from apply_workflow.schemas import ApplicationInput

logger = logging.getLogger(__name__)

# Keys are normalised (lowercase, stripped) at runtime — so casing in Excel
# doesn't matter. Values are ApplicationRow field names.
COLUMN_MAP = {
    "hr linkdin":    "hr_linkedin",   # covers 'HR LINkdin', 'hr linkdin', etc.
    "company":       "company",
    "post":          "post",
    "description":   "description",
    "email":         "to_email",
    "keywords":      "keywords",
    "spontane":      "spontane",
}


def _normalise(header: str) -> str:
    """Lowercase + strip — makes matching case/space insensitive."""
    return header.lower().strip()


def _require_file(excel_path: str) -> Path:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    return path


def _read_headers(ws) -> list[str]:
    """Return the sheet's header row, normalised."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [_normalise(str(h)) if h else "" for h in header_row]


def _field_to_col(norm_headers: list[str]) -> dict[str, int]:
    """Map ApplicationInput field name -> 1-based column index, using the
    first matching header for each field."""
    field_to_col: dict[str, int] = {}
    for norm_h, field in COLUMN_MAP.items():
        if field not in field_to_col and norm_h in norm_headers:
            field_to_col[field] = norm_headers.index(norm_h) + 1
    return field_to_col


def _check_row_in_range(ws, excel_row: int) -> None:
    if excel_row <= 1:
        raise ValueError("Cannot modify header row")
    if excel_row > ws.max_row:
        raise IndexError(f"Row {excel_row} is out of range")


def _read_application_records(excel_path: str) -> list[tuple[int, ApplicationInput]]:
    """Internal helper returning (excel_row_index, ApplicationInput) pairs."""
    path = _require_file(excel_path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.warning("Excel file is empty.")
        wb.close()
        return []

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    logger.info("Excel headers detected: %s", raw_headers)

    norm_index: dict[str, int] = {
        _normalise(h): i for i, h in enumerate(raw_headers) if h
    }
    logger.debug("Normalised header index: %s", norm_index)

    for norm_h in norm_index:
        if norm_h not in COLUMN_MAP:
            logger.warning("Unrecognised column (ignored): '%s'", norm_h)

    def _get(row_values: tuple, col_name: str) -> str:
        idx = norm_index.get(col_name)
        if idx is None or idx >= len(row_values):
            return ""
        val = row_values[idx]
        return str(val).strip() if val is not None else ""

    records: list[tuple[int, ApplicationInput]] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        company = _get(row, "company")
        if not company:
            logger.debug("Row %d skipped (no company).", row_idx)
            continue

        # Multiple headers can map to the same field; prefer the first
        # non-empty value found for each field.
        deduped: dict[str, str] = {}
        for norm_col, field in COLUMN_MAP.items():
            val = _get(row, norm_col)
            if val or field not in deduped:
                deduped[field] = val

        try:
            app = ApplicationInput(**deduped)
            records.append((row_idx, app))
            logger.info("Loaded row %d → %s / %s", row_idx, app.company, app.post)
        except Exception as e:
            logger.warning("Row %d skipped (validation error): %s", row_idx, e)

    wb.close()
    logger.info("Total applications loaded: %d", len(records))
    return records


def read_applications(excel_path: str) -> list[ApplicationInput]:
    """
    Parse the Excel file and return a list of ApplicationRow objects.
    Rows where 'company' is empty are skipped.
    """
    return [app for _, app in _read_application_records(excel_path)]


def read_applications_with_rows(excel_path: str) -> list[dict]:
    """Return application rows along with their Excel row indices."""
    return [
        {"excel_row": row_idx, **app.model_dump()}
        for row_idx, app in _read_application_records(excel_path)
    ]


def append_application(excel_path: str, data: dict) -> None:
    """Append one row to the Excel file, matching existing header order/casing."""
    path = _require_file(excel_path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    norm_headers = _read_headers(ws)
    field_to_col = _field_to_col(norm_headers)

    row_values = [""] * len(norm_headers)
    for field, value in data.items():
        col_idx = field_to_col.get(field)
        if col_idx is not None:
            row_values[col_idx - 1] = value

    ws.append(row_values)
    wb.save(path)
    logger.info("Appended new application row → %s", data.get("company"))


def update_application(excel_path: str, excel_row: int, data: dict) -> None:
    """Update an existing row in the Excel sheet by its 1-based Excel row index."""
    path = _require_file(excel_path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    _check_row_in_range(ws, excel_row)

    field_to_col = _field_to_col(_read_headers(ws))

    for field, value in data.items():
        col_idx = field_to_col.get(field)
        if col_idx is not None:
            ws.cell(row=excel_row, column=col_idx, value=value)

    wb.save(path)
    logger.info("Updated application row %d → %s", excel_row, data.get("company"))


def delete_application(excel_path: str, excel_row: int) -> None:
    """Delete an application row from the Excel sheet by its 1-based Excel row index."""
    path = _require_file(excel_path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    _check_row_in_range(ws, excel_row)

    field_to_col = _field_to_col(_read_headers(ws))
    company_col = field_to_col.get("company")
    company_val = ws.cell(row=excel_row, column=company_col).value if company_col else None

    ws.delete_rows(excel_row)
    wb.save(path)
    logger.info("Deleted application row %d → %s", excel_row, company_val or "?")